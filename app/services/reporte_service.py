from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

class ReporteService:
    """Servicio para generación de reportes en PDF y Excel"""

    @staticmethod
    def generar_pdf_resumen_aprendiz(aprendiz, matricula, output_path=None):
        """Genera PDF con resumen de matrícula del aprendiz"""
        if not output_path:
            # Usar ruta absoluta desde el directorio de la aplicación
            import tempfile
            temp_dir = tempfile.gettempdir()
            output_path = os.path.join(temp_dir, f"resumen_{aprendiz.usuario.documento}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

        pdf_doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("RESUMEN DE MATRÍCULA", title_style))
        story.append(Spacer(1, 0.3 * inch))

        # Datos del aprendiz
        story.append(Paragraph("<b>DATOS DEL APRENDIZ</b>", styles['Heading2']))
        datos_aprendiz = [
            ['Documento:', aprendiz.usuario.documento],
            ['Nombre Completo:', aprendiz.usuario.nombre_completo],
            ['Email:', aprendiz.usuario.email],
            ['Teléfono:', aprendiz.usuario.telefono or 'N/A'],
            ['Ciudad:', aprendiz.ciudad or 'N/A'],
        ]

        table = Table(datos_aprendiz, colWidths=[2 * inch, 4 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # Datos del acudiente
        story.append(Paragraph("<b>DATOS DEL ACUDIENTE</b>", styles['Heading2']))
        datos_acudiente = [
            ['Documento:', aprendiz.acudiente_documento or 'N/A'],
            ['Nombre Completo:', aprendiz.acudiente_nombre_completo or 'N/A'],
            ['Teléfono:', aprendiz.acudiente_telefono or 'N/A'],
            ['Email:', aprendiz.acudiente_email or 'N/A'],
        ]

        table = Table(datos_acudiente, colWidths=[2 * inch, 4 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # Datos académicos
        story.append(Paragraph("<b>DATOS ACADÉMICOS</b>", styles['Heading2']))
        datos_academicos = [
            ['Colegio:', aprendiz.colegio.nombre if aprendiz.colegio else 'N/A'],
            ['Grupo:', aprendiz.grupo.nombre if aprendiz.grupo else 'N/A'],
            ['Programa:', aprendiz.programa.nombre if aprendiz.programa else 'N/A'],
        ]

        table = Table(datos_academicos, colWidths=[2 * inch, 4 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # Documentos cargados
        story.append(Paragraph("<b>DOCUMENTOS CARGADOS</b>", styles['Heading2']))
        documentos_data = [['Tipo de Documento', 'Nombre Archivo', 'Validado']]

        for documento in matricula.documentos:
            if not documento.reemplazado_por:
                documentos_data.append([
                    documento.tipo_label,
                    documento.nombre_archivo[:40] + '...' if len(documento.nombre_archivo) > 40 else documento.nombre_archivo,
                    'Sí' if documento.validado else 'No'
                ])

        table = Table(documentos_data, colWidths=[2.5 * inch, 2.5 * inch, 1 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F8E9')]),
        ]))
        story.append(table)

        # Generar PDF
        pdf_doc.build(story)
        return output_path

    @staticmethod
    def generar_excel_matriculas(matriculas, output_path=None):
        """Genera Excel con listado de matrículas"""
        if not output_path:
            # Usar directorio temporal para evitar problemas de permisos
            import tempfile
            temp_dir = tempfile.gettempdir()
            output_path = os.path.join(temp_dir, f"matriculas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Matrículas"

        # Estilos
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_aligned = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = [
            'Documento', 'Nombres', 'Apellidos', 'Email', 'Teléfono',
            'Colegio', 'Grupo', 'Programa', 'Estado', 'Fecha Envío'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned

        # Datos
        for row, matricula in enumerate(matriculas, 2):
            aprendiz = matricula.aprendiz
            usuario = aprendiz.usuario

            ws.cell(row=row, column=1, value=usuario.documento)
            ws.cell(row=row, column=2, value=usuario.nombres)
            ws.cell(row=row, column=3, value=usuario.apellidos)
            ws.cell(row=row, column=4, value=usuario.email)
            ws.cell(row=row, column=5, value=usuario.telefono)
            ws.cell(row=row, column=6, value=aprendiz.colegio.nombre if aprendiz.colegio else 'N/A')
            ws.cell(row=row, column=7, value=aprendiz.grupo.nombre if aprendiz.grupo else 'N/A')
            ws.cell(row=row, column=8, value=aprendiz.programa.nombre if aprendiz.programa else 'N/A')
            ws.cell(row=row, column=9, value=matricula.estado)
            ws.cell(row=row, column=10, value=matricula.fecha_envio.strftime('%Y-%m-%d') if matricula.fecha_envio else 'N/A')

        # Ajustar anchos de columna
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        wb.save(output_path)
        return output_path

    @staticmethod
    def generar_pdf_reporte_docente(matriculas, titulo, colegio_nombre, filtros=None, output_path=None):
        """Genera PDF con reporte de matrículas para docente"""
        if not output_path:
            import tempfile
            temp_dir = tempfile.gettempdir()
            output_path = os.path.join(temp_dir, f"reporte_docente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

        pdf_doc = SimpleDocTemplate(output_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # Título principal
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        story.append(Paragraph(titulo, title_style))

        # Subtítulo con colegio
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#4a5568'),
            spaceAfter=10,
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"<b>Colegio:</b> {colegio_nombre}", subtitle_style))
        story.append(Paragraph(f"<b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
        story.append(Spacer(1, 0.2 * inch))

        # Filtros aplicados
        if filtros:
            story.append(Paragraph("<b>Filtros Aplicados:</b>", styles['Heading3']))
            filtros_text = "<br/>".join([f"• {k}: {v}" for k, v in filtros.items() if v])
            story.append(Paragraph(filtros_text, styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))

        # Estadísticas
        stats = {
            'total': len(matriculas),
            'borrador': len([m for m in matriculas if m.estado == 'BORRADOR']),
            'enviado': len([m for m in matriculas if m.estado == 'ENVIADO']),
            'prematricula': len([m for m in matriculas if m.estado == 'PREMATRICULA']),
            'matriculado': len([m for m in matriculas if m.estado == 'MATRICULADO'])
        }

        story.append(Paragraph("<b>RESUMEN ESTADÍSTICO</b>", styles['Heading2']))
        stats_data = [
            ['Estado', 'Cantidad'],
            ['Total Matrículas', str(stats['total'])],
            ['Borrador', str(stats['borrador'])],
            ['Enviado', str(stats['enviado'])],
            ['Pre-matrícula', str(stats['prematricula'])],
            ['Matriculado', str(stats['matriculado'])]
        ]

        table = Table(stats_data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # Listado de matrículas
        if matriculas:
            story.append(Paragraph("<b>LISTADO DE MATRÍCULAS</b>", styles['Heading2']))

            matriculas_data = [['Documento', 'Aprendiz', 'Programa', 'Grupo', 'Estado']]

            for matricula in matriculas:
                aprendiz = matricula.aprendiz
                usuario = aprendiz.usuario

                # Truncar nombres largos
                nombre = usuario.nombre_completo
                if len(nombre) > 30:
                    nombre = nombre[:27] + '...'

                programa = aprendiz.programa.nombre[:25] if aprendiz.programa else 'N/A'
                grupo = aprendiz.grupo.nombre if aprendiz.grupo else 'N/A'

                matriculas_data.append([
                    usuario.documento,
                    nombre,
                    programa,
                    grupo,
                    matricula.estado
                ])

            table = Table(matriculas_data, colWidths=[1.2*inch, 2*inch, 2*inch, 1*inch, 1.3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
            ]))
            story.append(table)

        # Pie de página
        story.append(Spacer(1, 0.5 * inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"Reporte generado por Sistema de Matrículas SENA - {datetime.now().strftime('%d/%m/%Y %H:%M')}", footer_style))

        # Generar PDF
        pdf_doc.build(story)
        return output_path
