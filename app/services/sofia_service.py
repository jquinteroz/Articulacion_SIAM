from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from app.models import Aprendiz, Grupo, Programa, Colegio
from sqlalchemy.orm import joinedload


class SofiaService:
    """Servicio para generar formatos SOFIA Plus"""

    @staticmethod
    def generar_formato_sofia(filtro_tipo, filtro_id=None, docente_colegio_id=None):
        """
        Genera formato SOFIA Plus en Excel con los aprendices filtrados

        Args:
            filtro_tipo: 'colegio', 'ficha', 'programa', 'todos'
            filtro_id: ID del colegio, ficha o programa
            docente_colegio_id: ID del colegio del docente (para validar permisos)

        Returns:
            tuple: (success, message, file_path)
        """
        try:
            # Usar el nuevo template con formato correcto
            template_path = os.path.join('formatos', 'Sofia new.xlsx')

            if not os.path.exists(template_path):
                return False, 'No se encontró el template del formato SOFIA', None

            # Cargar el workbook del template (esto preserva todos los formatos)
            wb = load_workbook(template_path)
            ws = wb.active

            # Obtener aprendices según filtro
            query = Aprendiz.query.join(Aprendiz.usuario).join(Aprendiz.grupo)

            # Validar permisos de docente
            if docente_colegio_id:
                # Si es docente, solo puede ver su colegio
                if filtro_tipo == 'colegio' and filtro_id != docente_colegio_id:
                    return False, 'No tiene permisos para generar este reporte', None
                elif filtro_tipo == 'ficha':
                    # Verificar que la ficha pertenezca a su colegio
                    grupo = Grupo.query.get(filtro_id)
                    if not grupo or grupo.colegio_id != docente_colegio_id:
                        return False, 'No tiene permisos para esta ficha', None

                # Limitar a su colegio
                query = query.filter(Aprendiz.colegio_id == docente_colegio_id)

            # Aplicar filtros
            if filtro_tipo == 'colegio' and filtro_id:
                query = query.filter(Aprendiz.colegio_id == filtro_id)
                filtro_nombre = Colegio.query.get(filtro_id).nombre if filtro_id else 'Todos'
            elif filtro_tipo == 'ficha' and filtro_id:
                query = query.filter(Aprendiz.grupo_id == filtro_id)
                grupo = Grupo.query.get(filtro_id)
                filtro_nombre = grupo.nombre if grupo else 'Todos'
            elif filtro_tipo == 'programa' and filtro_id:
                query = query.filter(Aprendiz.programa_id == filtro_id)
                programa = Programa.query.get(filtro_id)
                filtro_nombre = programa.nombre if programa else 'Todos'
            else:
                filtro_nombre = 'Todos'

            aprendices = query.all()

            if not aprendices:
                return False, 'No se encontraron aprendices con los filtros seleccionados', None

            # Escribir datos de aprendices a partir de la fila 3 (row 3 en openpyxl, indexado desde 1)
            # Asumiendo que fila 1 es título y fila 2 son encabezados
            row_idx = 3

            for aprendiz in aprendices:
                usuario = aprendiz.usuario
                grupo = aprendiz.grupo

                # Columna A (1): Resultado del Registro (vacío - reservado para sistema)
                ws.cell(row=row_idx, column=1).value = ''

                # Columna B (2): Tipo de Identificación
                tipo_doc = usuario.tipo_documento if usuario.tipo_documento else 'CC'
                ws.cell(row=row_idx, column=2).value = tipo_doc

                # Columna C (3): Número de Identificación
                ws.cell(row=row_idx, column=3).value = str(usuario.documento)

                # Columna D (4): Código de la ficha (número del grupo)
                codigo_ficha = grupo.nombre if grupo else ''
                ws.cell(row=row_idx, column=4).value = codigo_ficha

                # Columna E (5): Tipo Población Aspirante (siempre "NINGUNA")
                ws.cell(row=row_idx, column=5).value = 'NINGUNA'

                # Columnas F y G (6-7): Vacías
                ws.cell(row=row_idx, column=6).value = ''
                ws.cell(row=row_idx, column=7).value = ''

                # Columnas H-I (8-9): Mantener metadata del template (no tocar)
                # openpyxl preserva automáticamente el contenido y formato de estas columnas

                row_idx += 1

            # Generar nombre de archivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'SOFIA_{filtro_tipo}_{filtro_nombre.replace(" ", "_")}_{timestamp}.xlsx'

            # Usar ruta absoluta al directorio temp del proyecto
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            temp_dir = os.path.join(base_dir, 'temp')

            # Crear directorio temp si no existe
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)

            output_path = os.path.join(temp_dir, filename)

            # Guardar archivo (openpyxl preserva todos los formatos del template)
            wb.save(output_path)

            return True, f'Formato SOFIA generado con {len(aprendices)} aprendices', output_path

        except Exception as e:
            return False, f'Error al generar formato SOFIA: {str(e)}', None

    @staticmethod
    def get_opciones_filtro_admin():
        """Obtiene las opciones de filtro para administrador"""
        colegios = Colegio.query.filter_by(activo=True).order_by(Colegio.nombre).all()
        grupos = Grupo.query.order_by(Grupo.nombre).all()
        programas = Programa.query.filter_by(activo=True).order_by(Programa.nombre).all()

        return {
            'colegios': colegios,
            'grupos': grupos,
            'programas': programas
        }

    @staticmethod
    def get_opciones_filtro_docente(colegio_id):
        """Obtiene las opciones de filtro para docente enlace"""
        colegio = Colegio.query.get(colegio_id)
        grupos = Grupo.query.options(joinedload(Grupo.programa)).filter_by(colegio_id=colegio_id).order_by(Grupo.nombre).all()

        return {
            'colegio': colegio,
            'grupos': grupos
        }
