from flask import render_template, redirect, url_for, flash, request, send_file, jsonify, current_app
from flask_login import login_required, current_user
from app.models import db, Matricula, Aprendiz, Colegio, Grupo, Programa, Documento, Usuario, DocumentoSIMAT
from app.services.matricula_service import MatriculaService
from app.services.documento_service import DocumentoService
from app.services.reporte_service import ReporteService
from app.services.sofia_service import SofiaService
from app.utils.decorators import docente_required
from app.utils.crypto import CryptoService
from . import docente_bp
from datetime import datetime
from werkzeug.utils import secure_filename
import os

@docente_bp.route('/dashboard')
@login_required
@docente_required
def dashboard():
    """Dashboard del docente enlace"""
    # Obtener colegio del docente
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    # Inicializar estadísticas
    stats = {
        'total_matriculas': 0,
        'matriculas_pendientes': 0,
        'matriculas_validadas': 0,
        'documentos_pendientes': 0
    }

    matriculas_recientes = []

    if colegio:
        # Obtener todas las matrículas del colegio
        matriculas = Matricula.query.join(Aprendiz).filter(
            Aprendiz.colegio_id == colegio.id
        ).all()

        stats['total_matriculas'] = len(matriculas)
        stats['matriculas_pendientes'] = len([m for m in matriculas if m.estado in ['ENVIADO', 'PENDIENTE']])
        stats['matriculas_validadas'] = len([m for m in matriculas if m.estado in ['COMPLETO', 'PREMATRICULA']])

        # Contar documentos pendientes
        for matricula in matriculas:
            stats['documentos_pendientes'] += len([d for d in matricula.documentos if d.estado == 'PENDIENTE'])

        # Matrículas recientes (últimas 5)
        matriculas_recientes = Matricula.query.join(Aprendiz).filter(
            Aprendiz.colegio_id == colegio.id
        ).order_by(Matricula.created_at.desc()).limit(5).all()

    return render_template('docente/dashboard.html',
                         colegio=colegio,
                         stats=stats,
                         matriculas_recientes=matriculas_recientes)

@docente_bp.route('/matriculas')
@login_required
@docente_required
def matriculas():
    """Listado de matrículas"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No tiene un colegio asignado', 'warning')
        return redirect(url_for('docente.dashboard'))

    # Filtros
    estado_filtro = request.args.get('estado', 'todos')
    grupo_filtro = request.args.get('grupo', 'todos')

    # Query con eager loading para evitar N+1 queries
    query = Matricula.query.join(Aprendiz).filter(Aprendiz.colegio_id == colegio.id)\
        .options(
            db.joinedload(Matricula.aprendiz).joinedload(Aprendiz.usuario),
            db.joinedload(Matricula.aprendiz).joinedload(Aprendiz.grupo),
            db.joinedload(Matricula.aprendiz).joinedload(Aprendiz.programa),
            db.joinedload(Matricula.documentos)
        )

    if estado_filtro != 'todos':
        query = query.filter(Matricula.estado == estado_filtro)

    if grupo_filtro != 'todos':
        query = query.filter(Aprendiz.grupo_id == int(grupo_filtro))

    matriculas = query.order_by(Matricula.created_at.desc()).all()
    grupos = Grupo.query.filter_by(colegio_id=colegio.id, activo=True).all()

    # Debug: Log para verificar qué se está cargando
    current_app.logger.info(f"Estado filtro: {estado_filtro}, Grupo filtro: {grupo_filtro}")
    current_app.logger.info(f"Total matrículas encontradas: {len(matriculas)}")

    # Flash temporal para debug
    if estado_filtro == 'PENDIENTE':
        flash(f'Buscando matrículas con estado PENDIENTE. Se encontraron: {len(matriculas)}', 'info')

    if matriculas:
        for m in matriculas:
            aprendiz_nombre = 'N/A'
            try:
                if m.aprendiz and m.aprendiz.usuario:
                    aprendiz_nombre = m.aprendiz.usuario.nombre_completo
            except Exception as e:
                current_app.logger.error(f"Error al obtener nombre del aprendiz: {e}")
            current_app.logger.info(f"Matrícula ID: {m.id}, Estado: {m.estado}, Aprendiz: {aprendiz_nombre}")
    else:
        current_app.logger.warning("No se encontraron matrículas con los filtros aplicados")
        if estado_filtro != 'todos':
            flash(f'No se encontraron matrículas con el filtro de estado: {estado_filtro}', 'warning')

    # Debug adicional: Mostrar todos los estados disponibles
    if estado_filtro == 'PENDIENTE' and len(matriculas) == 0:
        todas_matriculas = Matricula.query.join(Aprendiz).filter(Aprendiz.colegio_id == colegio.id).all()
        estados_disponibles = set([m.estado for m in todas_matriculas])
        current_app.logger.info(f"Estados disponibles en la BD: {estados_disponibles}")
        flash(f'Estados disponibles en la base de datos: {", ".join(estados_disponibles)}', 'info')

    return render_template('docente/matriculas.html',
                         matriculas=matriculas,
                         grupos=grupos,
                         estado_filtro=estado_filtro,
                         grupo_filtro=grupo_filtro)

@docente_bp.route('/matricula/<int:matricula_id>')
@login_required
@docente_required
def ver_matricula(matricula_id):
    """Ver detalle de una matrícula"""
    matricula = Matricula.query.get_or_404(matricula_id)
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    # Verificar que la matrícula pertenece al colegio del docente
    if matricula.aprendiz.colegio_id != colegio.id:
        flash('No tiene permisos para ver esta matrícula', 'danger')
        return redirect(url_for('docente.matriculas'))

    documentos = DocumentoService.get_documentos_activos(matricula.id)

    # Cargar grupos y programas para edición
    grupos = Grupo.query.filter_by(activo=True, colegio_id=colegio.id).options(
        db.joinedload(Grupo.programa)
    ).all()
    programas = Programa.query.filter_by(activo=True).all()

    return render_template('docente/ver_matricula.html',
                         matricula=matricula,
                         documentos=documentos,
                         grupos=grupos,
                         programas=programas)

@docente_bp.route('/editar-datos-aprendiz/<int:matricula_id>', methods=['POST'])
@login_required
@docente_required
def editar_datos_aprendiz(matricula_id):
    """Editar grupo y programa del aprendiz"""
    matricula = Matricula.query.get_or_404(matricula_id)
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    # Verificar que la matrícula pertenece al colegio del docente
    if matricula.aprendiz.colegio_id != colegio.id:
        flash('No tiene permisos para editar esta matrícula', 'danger')
        return redirect(url_for('docente.matriculas'))

    try:
        aprendiz = matricula.aprendiz

        # Obtener datos del formulario
        tipo_documento = request.form.get('tipo_documento')
        grupo_id = request.form.get('grupo_id')
        programa_id = request.form.get('programa_id')

        # Actualizar tipo de documento
        if tipo_documento and tipo_documento in ['TI', 'CC', 'CE', 'PEP', 'PPT']:
            aprendiz.usuario.tipo_documento = tipo_documento

        # Validar que el grupo pertenece al colegio del docente
        if grupo_id:
            grupo = Grupo.query.get(grupo_id)
            if not grupo or grupo.colegio_id != colegio.id:
                flash('El grupo seleccionado no es válido para este colegio', 'danger')
                return redirect(url_for('docente.ver_matricula', matricula_id=matricula_id))

            aprendiz.grupo_id = grupo_id

        # Actualizar programa
        if programa_id:
            aprendiz.programa_id = programa_id

        db.session.commit()
        flash('Datos del aprendiz actualizados exitosamente', 'success')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al editar datos del aprendiz: {e}")
        flash(f'Error al actualizar datos: {str(e)}', 'danger')

    return redirect(url_for('docente.ver_matricula', matricula_id=matricula_id))

@docente_bp.route('/validar-matricula/<int:matricula_id>', methods=['POST'])
@login_required
@docente_required
def validar_matricula(matricula_id):
    """Validar o rechazar una matrícula"""
    matricula = Matricula.query.get_or_404(matricula_id)
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if matricula.aprendiz.colegio_id != colegio.id:
        return jsonify({'success': False, 'message': 'No tiene permisos'}), 403

    estado = request.form.get('estado')
    observaciones = request.form.get('observaciones')

    success, message = MatriculaService.validar_matricula_docente(
        matricula_id,
        current_user.id,
        estado,
        observaciones
    )

    if success:
        flash('Matrícula validada exitosamente', 'success')
    else:
        flash(message, 'danger')

    return redirect(url_for('docente.ver_matricula', matricula_id=matricula_id))

@docente_bp.route('/reemplazar-documento/<int:documento_id>', methods=['POST'])
@login_required
@docente_required
def reemplazar_documento(documento_id):
    """Reemplazar un documento"""
    documento = Documento.query.get_or_404(documento_id)
    file = request.files.get('archivo')

    if not file or not file.filename:
        flash('Debe seleccionar un archivo', 'danger')
        return redirect(request.referrer)

    success, message, nuevo_doc = DocumentoService.replace_documento(
        documento_id,
        file,
        current_user.id
    )

    if success:
        flash('Documento reemplazado exitosamente', 'success')
    else:
        flash(message, 'danger')

    return redirect(request.referrer)

@docente_bp.route('/reportes')
@login_required
@docente_required
def reportes():
    """Página de reportes"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()
    grupos = Grupo.query.filter_by(colegio_id=colegio.id, activo=True).all() if colegio else []
    programas = Programa.query.filter_by(activo=True).all()

    # Obtener estadísticas de matrículas
    matriculas = Matricula.query.join(Aprendiz).filter(Aprendiz.colegio_id == colegio.id).all() if colegio else []

    stats = {
        'total': len(matriculas),
        'pendientes': len([m for m in matriculas if m.estado in ['ENVIADO', 'PENDIENTE']]),
        'completas': len([m for m in matriculas if m.estado == 'COMPLETO']),
        'prematriculas': len([m for m in matriculas if m.estado == 'PREMATRICULA'])
    }

    return render_template('docente/reportes.html',
                         colegio=colegio,
                         grupos=grupos,
                         programas=programas,
                         stats=stats)

@docente_bp.route('/generar-reporte-pdf', methods=['POST'])
@login_required
@docente_required
def generar_reporte_pdf():
    """Generar reporte en PDF"""
    from flask import send_file
    from app.services.reporte_service import ReporteService
    from datetime import datetime

    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No tiene un colegio asignado', 'warning')
        return redirect(url_for('docente.dashboard'))

    # Obtener filtros
    filtro_tipo = request.form.get('filtro_tipo')
    programa_id = request.form.get('programa_id')
    grupo_id = request.form.get('grupo_id')
    estado = request.form.get('estado')
    fecha_desde = request.form.get('fecha_desde')
    fecha_hasta = request.form.get('fecha_hasta')

    # Query base: todas las matrículas del colegio
    query = Matricula.query.join(Aprendiz).filter(Aprendiz.colegio_id == colegio.id)

    # Aplicar filtros
    filtros = {}

    if programa_id:
        query = query.filter(Aprendiz.programa_id == int(programa_id))
        programa = Programa.query.get(programa_id)
        filtros['Programa'] = programa.nombre if programa else 'N/A'

    if grupo_id:
        query = query.filter(Aprendiz.grupo_id == int(grupo_id))
        grupo = Grupo.query.get(grupo_id)
        filtros['Grupo'] = grupo.nombre if grupo else 'N/A'

    if estado:
        query = query.filter(Matricula.estado == estado)
        filtros['Estado'] = estado

    if fecha_desde:
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
        query = query.filter(Matricula.created_at >= fecha_desde_dt)
        filtros['Fecha Desde'] = fecha_desde

    if fecha_hasta:
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        query = query.filter(Matricula.created_at <= fecha_hasta_dt)
        filtros['Fecha Hasta'] = fecha_hasta

    matriculas = query.all()

    # Determinar título según tipo de reporte
    titulos = {
        'matriculas': 'Reporte de Matrículas por Estado',
        'programa': 'Reporte de Matrículas por Programa',
        'grupo': 'Reporte de Matrículas por Grupo',
        'documentos': 'Reporte de Estado de Documentos',
        'completo': 'Reporte Completo de Matrículas'
    }
    titulo = titulos.get(filtro_tipo, 'Reporte de Matrículas')

    try:
        pdf_path = ReporteService.generar_pdf_reporte_docente(
            matriculas,
            titulo,
            colegio.nombre,
            filtros
        )
        return send_file(pdf_path, as_attachment=True, download_name=f'reporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf')

    except Exception as e:
        flash(f'Error al generar reporte PDF: {str(e)}', 'danger')
        return redirect(url_for('docente.reportes'))

@docente_bp.route('/generar-reporte-excel', methods=['POST'])
@login_required
@docente_required
def generar_reporte_excel():
    """Generar reporte en Excel"""
    from flask import send_file
    from app.services.reporte_service import ReporteService
    from datetime import datetime

    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No tiene un colegio asignado', 'warning')
        return redirect(url_for('docente.dashboard'))

    # Obtener tipo de reporte
    filtro_tipo = request.form.get('filtro_tipo', 'completo')

    # Si es formato SOFIA, usar el servicio SOFIA
    if filtro_tipo == 'sofia':
        try:
            grupo_id = request.form.get('grupo_id')

            # Determinar filtro para SOFIA
            if grupo_id:
                tipo_filtro = 'ficha'
                id_filtro = int(grupo_id)
            else:
                tipo_filtro = 'colegio'
                id_filtro = colegio.id

            # Generar formato SOFIA
            success, message, file_path = SofiaService.generar_formato_sofia(
                filtro_tipo=tipo_filtro,
                filtro_id=id_filtro,
                docente_colegio_id=colegio.id
            )

            if success:
                return send_file(
                    file_path,
                    mimetype='application/vnd.ms-excel',
                    as_attachment=True,
                    download_name=os.path.basename(file_path)
                )
            else:
                flash(message, 'danger')
                return redirect(url_for('docente.reportes'))

        except Exception as e:
            flash(f'Error al generar formato SOFIA: {str(e)}', 'danger')
            return redirect(url_for('docente.reportes'))

    # Para otros tipos de reportes, usar el servicio de reportes normal
    # Obtener filtros
    programa_id = request.form.get('programa_id')
    grupo_id = request.form.get('grupo_id')
    estado = request.form.get('estado')
    fecha_desde = request.form.get('fecha_desde')
    fecha_hasta = request.form.get('fecha_hasta')

    # Query base: todas las matrículas del colegio
    query = Matricula.query.join(Aprendiz).filter(Aprendiz.colegio_id == colegio.id)

    # Aplicar filtros
    if programa_id:
        query = query.filter(Aprendiz.programa_id == int(programa_id))

    if grupo_id:
        query = query.filter(Aprendiz.grupo_id == int(grupo_id))

    if estado:
        query = query.filter(Matricula.estado == estado)

    if fecha_desde:
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
        query = query.filter(Matricula.created_at >= fecha_desde_dt)

    if fecha_hasta:
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        query = query.filter(Matricula.created_at <= fecha_hasta_dt)

    matriculas = query.all()

    try:
        excel_path = ReporteService.generar_excel_matriculas(matriculas)
        return send_file(excel_path, as_attachment=True, download_name=f'reporte_matriculas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

    except Exception as e:
        flash(f'Error al generar reporte Excel: {str(e)}', 'danger')
        return redirect(url_for('docente.reportes'))

# ============================================
# CRUD DE APRENDICES
# ============================================

@docente_bp.route('/aprendices')
@login_required
@docente_required
def aprendices():
    """Lista de aprendices del colegio"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No tiene un colegio asignado', 'warning')
        return redirect(url_for('docente.dashboard'))

    # Filtros
    nombre_filtro = request.args.get('nombre', '')
    programa_filtro = request.args.get('programa', '')
    grupo_filtro = request.args.get('grupo', '')
    estado_filtro = request.args.get('estado', '')

    # Query base
    query = Aprendiz.query.join(Usuario).filter(Aprendiz.colegio_id == colegio.id)

    # Aplicar filtros
    if nombre_filtro:
        query = query.filter(
            db.or_(
                Usuario.nombres.ilike(f'%{nombre_filtro}%'),
                Usuario.apellidos.ilike(f'%{nombre_filtro}%'),
                Usuario.documento.ilike(f'%{nombre_filtro}%')
            )
        )

    if programa_filtro:
        query = query.join(Matricula).filter(Matricula.programa_id == programa_filtro)

    if grupo_filtro:
        query = query.filter(Aprendiz.grupo_id == grupo_filtro)

    if estado_filtro:
        query = query.join(Matricula).filter(Matricula.estado == estado_filtro)

    aprendices = query.all()

    # Datos para filtros
    programas = Programa.query.filter_by(activo=True).all()
    grupos = Grupo.query.filter_by(colegio_id=colegio.id, activo=True).all()

    # Estadísticas
    stats = {
        'total': Aprendiz.query.filter_by(colegio_id=colegio.id).count(),
        'con_matricula': Aprendiz.query.join(Matricula).filter(Aprendiz.colegio_id == colegio.id).count(),
        'sin_matricula': Aprendiz.query.filter_by(colegio_id=colegio.id).filter(~Aprendiz.matriculas.any()).count(),
        'activos': Aprendiz.query.join(Usuario).filter(Aprendiz.colegio_id == colegio.id, Usuario.activo == True).count()
    }

    return render_template('docente/aprendices/list.html',
                         aprendices=aprendices,
                         colegio=colegio,
                         programas=programas,
                         grupos=grupos,
                         stats=stats,
                         nombre_filtro=nombre_filtro,
                         programa_filtro=programa_filtro,
                         grupo_filtro=grupo_filtro,
                         estado_filtro=estado_filtro)

@docente_bp.route('/aprendices/<int:aprendiz_id>')
@login_required
@docente_required
def ver_aprendiz(aprendiz_id):
    """Ver detalle de un aprendiz"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()
    aprendiz = Aprendiz.query.filter_by(id=aprendiz_id, colegio_id=colegio.id).first_or_404()

    # Obtener matrículas del aprendiz
    matriculas = Matricula.query.filter_by(aprendiz_id=aprendiz.id).order_by(Matricula.created_at.desc()).all()

    return render_template('docente/aprendices/detalle.html',
                         aprendiz=aprendiz,
                         matriculas=matriculas,
                         colegio=colegio)

@docente_bp.route('/aprendices/plantilla-excel')
@login_required
@docente_required
def descargar_plantilla_excel():
    """Descarga plantilla Excel para carga masiva de aprendices"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Aprendices"

    # Encabezados
    headers = ['documento', 'tipo_documento', 'nombres', 'apellidos', 'email', 'telefono', 'direccion', 'ciudad', 'departamento']
    ws.append(headers)

    # Estilo para encabezados
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fila de ejemplo
    ejemplo = ['1000000003', 'CC', 'Juan Carlos', 'Pérez López', 'estudiante@mail.com', '3001234567', 'Calle 10 #5-20', 'Bogotá', 'Cundinamarca']
    ws.append(ejemplo)

    # Ajustar anchos de columna
    column_widths = [15, 18, 20, 20, 30, 15, 25, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_aprendices.xlsx'
    )

@docente_bp.route('/aprendices/cargar-excel', methods=['GET', 'POST'])
@login_required
@docente_required
def cargar_excel_aprendices():
    """Carga masiva de aprendices desde Excel"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No tiene un colegio asignado', 'warning')
        return redirect(url_for('docente.dashboard'))

    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'danger')
            return redirect(url_for('docente.cargar_excel_aprendices'))

        archivo = request.files['archivo']

        if archivo.filename == '':
            flash('No se seleccionó ningún archivo', 'danger')
            return redirect(url_for('docente.cargar_excel_aprendices'))

        if not archivo.filename.endswith(('.xlsx', '.xls')):
            flash('El archivo debe ser Excel (.xlsx o .xls)', 'danger')
            return redirect(url_for('docente.cargar_excel_aprendices'))

        try:
            import pandas as pd
            from app.services.auth_service import AuthService

            # Leer Excel
            df = pd.read_excel(archivo)

            # Validar columnas requeridas
            columnas_requeridas = ['documento', 'tipo_documento', 'nombres', 'apellidos', 'email', 'telefono']
            if not all(col in df.columns for col in columnas_requeridas):
                flash(f'El archivo debe contener las columnas: {", ".join(columnas_requeridas)}', 'danger')
                return redirect(url_for('docente.cargar_excel_aprendices'))

            # Procesar cada fila
            exitosos = 0
            errores = []

            for index, row in df.iterrows():
                try:
                    # Crear usuario aprendiz
                    # Contraseña por defecto: documento
                    password = str(row['documento'])

                    success, message, usuario = AuthService.create_user(
                        documento=str(row['documento']),
                        tipo_documento=row['tipo_documento'],
                        nombres=row['nombres'],
                        apellidos=row['apellidos'],
                        email=row['email'],
                        telefono=str(row['telefono']) if pd.notna(row['telefono']) else '',
                        password=password,
                        rol='APRENDIZ'
                    )

                    if success:
                        # Vincular aprendiz al colegio
                        aprendiz = Aprendiz.query.filter_by(usuario_id=usuario.id).first()
                        if aprendiz:
                            aprendiz.colegio_id = colegio.id
                            aprendiz.ciudad = row.get('ciudad', '')
                            aprendiz.departamento = row.get('departamento', '')

                            db.session.commit()
                            exitosos += 1
                    else:
                        errores.append(f'Fila {index + 2}: {message}')

                except Exception as e:
                    errores.append(f'Fila {index + 2}: {str(e)}')
                    continue

            # Mostrar resultados
            if exitosos > 0:
                flash(f'Se importaron exitosamente {exitosos} aprendices', 'success')

            if errores:
                flash(f'Errores encontrados: {len(errores)}', 'warning')
                for error in errores[:5]:  # Mostrar solo los primeros 5 errores
                    flash(error, 'danger')

            return redirect(url_for('docente.aprendices'))

        except Exception as e:
            flash(f'Error al procesar el archivo: {str(e)}', 'danger')
            return redirect(url_for('docente.cargar_excel_aprendices'))

    # GET - Mostrar formulario de carga
    grupos = Grupo.query.filter_by(colegio_id=colegio.id, activo=True).all()
    return render_template('docente/aprendices/cargar_excel.html',
                         colegio=colegio,
                         grupos=grupos)

# ============================================
# VISUALIZACIÓN Y APROBACIÓN DE DOCUMENTOS
# ============================================

@docente_bp.route('/ver-documento/<int:documento_id>')
@login_required
@docente_required
def ver_documento(documento_id):
    """Ver un documento específico"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()
    documento = Documento.query.get_or_404(documento_id)

    # Verificar que el documento pertenece a un aprendiz del colegio
    if documento.matricula.aprendiz.colegio_id != colegio.id:
        flash('No tiene permisos para ver este documento', 'danger')
        return redirect(url_for('docente.dashboard'))

    # Convertir ruta relativa a absoluta si es necesario
    file_path = documento.ruta_archivo
    if not os.path.isabs(file_path):
        # Es una ruta relativa, convertirla a absoluta
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        file_path = os.path.join(base_dir, file_path)

    if not os.path.exists(file_path):
        flash(f'Archivo no encontrado: {os.path.basename(file_path)}', 'danger')
        return redirect(url_for('docente.ver_matricula', matricula_id=documento.matricula_id))

    # Detectar el tipo de archivo para usar el mimetype correcto
    extension = os.path.splitext(file_path)[1].lower()
    mimetype_map = {
        '.pdf': 'application/pdf',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg'
    }
    mimetype = mimetype_map.get(extension, 'application/pdf')

    return send_file(file_path, mimetype=mimetype)

@docente_bp.route('/aprobar-documento/<int:documento_id>', methods=['POST'])
@login_required
@docente_required
def aprobar_documento(documento_id):
    """Aprobar un documento"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()
    documento = Documento.query.get_or_404(documento_id)

    # Verificar que el documento pertenece a un aprendiz del colegio
    if documento.matricula.aprendiz.colegio_id != colegio.id:
        flash('No tiene permisos para aprobar este documento', 'danger')
        return redirect(url_for('docente.dashboard'))

    # Verificar que el documento está en estado pendiente de revisión
    if documento.estado not in ['PENDIENTE', 'ENVIADO']:
        flash('Este documento no puede ser aprobado en su estado actual', 'warning')
        return redirect(url_for('docente.ver_matricula', matricula_id=documento.matricula_id))

    # Aprobar documento usando los campos del modelo
    documento.validado = True
    documento.validado_por = current_user.id
    documento.fecha_validacion = datetime.utcnow()
    documento.observaciones = 'Documento aprobado'

    # Obtener matrícula antes de commit
    matricula = documento.matricula

    db.session.commit()

    # Refrescar documentos después del commit
    db.session.refresh(documento)

    flash(f'Documento {documento.tipo_documento.replace("_", " ").title()} aprobado exitosamente', 'success')

    # Verificar si todos los documentos están aprobados (refrescar desde DB)
    documentos_matricula = Documento.query.filter_by(matricula_id=matricula.id).all()
    documentos_activos = [d for d in documentos_matricula if not d.reemplazado_por]
    todos_aprobados = all(d.validado == True for d in documentos_activos)

    if todos_aprobados and len(documentos_activos) >= 8:
        matricula.estado = 'PREMATRICULA'
        matricula.fecha_validacion_docente = datetime.utcnow()
        matricula.validado_por_docente = current_user.id
        db.session.commit()
        flash('Todos los documentos aprobados. El aprendiz pasa a estado PRE-MATRÍCULA', 'success')

    return redirect(url_for('docente.ver_matricula', matricula_id=documento.matricula_id))

@docente_bp.route('/rechazar-documento/<int:documento_id>', methods=['POST'])
@login_required
@docente_required
def rechazar_documento(documento_id):
    """Rechazar un documento"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()
    documento = Documento.query.get_or_404(documento_id)

    # Verificar que el documento pertenece a un aprendiz del colegio
    if documento.matricula.aprendiz.colegio_id != colegio.id:
        flash('No tiene permisos para rechazar este documento', 'danger')
        return redirect(url_for('docente.dashboard'))

    observaciones = request.form.get('observaciones', '')

    if not observaciones:
        flash('Debe proporcionar observaciones al rechazar un documento', 'warning')
        return redirect(url_for('docente.ver_matricula', matricula_id=documento.matricula_id))

    # Rechazar documento usando los campos del modelo
    documento.validado = False
    documento.validado_por = current_user.id
    documento.fecha_validacion = datetime.utcnow()
    documento.observaciones = observaciones
    db.session.commit()

    flash(f'Documento {documento.tipo_documento.replace("_", " ").title()} rechazado. El aprendiz debe cargarlo nuevamente', 'warning')

    return redirect(url_for('docente.ver_matricula', matricula_id=documento.matricula_id))

@docente_bp.route('/aprobar-todos-documentos/<int:matricula_id>', methods=['POST'])
@login_required
@docente_required
def aprobar_todos_documentos(matricula_id):
    """Aprobar todos los documentos pendientes de una matrícula"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()
    matricula = Matricula.query.get_or_404(matricula_id)

    # Verificar que la matrícula pertenece a un aprendiz del colegio
    if matricula.aprendiz.colegio_id != colegio.id:
        flash('No tiene permisos para aprobar documentos de esta matrícula', 'danger')
        return redirect(url_for('docente.dashboard'))

    # Obtener todos los documentos pendientes o enviados
    documentos = Documento.query.filter_by(matricula_id=matricula_id).all()
    documentos_aprobados = 0

    for documento in documentos:
        if documento.estado in ['PENDIENTE', 'ENVIADO']:
            documento.validado = True
            documento.validado_por = current_user.id
            documento.fecha_validacion = datetime.utcnow()
            documento.observaciones = 'Documento aprobado'
            documentos_aprobados += 1

    db.session.commit()

    if documentos_aprobados > 0:
        flash(f'Se aprobaron {documentos_aprobados} documento(s) exitosamente', 'success')

        # Refrescar documentos desde la base de datos
        documentos = Documento.query.filter_by(matricula_id=matricula_id).all()

        # Verificar si todos los documentos están aprobados usando validado
        todos_aprobados = all(d.validado == True and d.reemplazado_por is None for d in documentos)

        # Contar solo documentos activos (no reemplazados)
        documentos_activos = [d for d in documentos if not d.reemplazado_por]

        if todos_aprobados and len(documentos_activos) >= 8:
            matricula.estado = 'PREMATRICULA'
            matricula.fecha_validacion_docente = datetime.utcnow()
            matricula.validado_por_docente = current_user.id
            db.session.commit()
            flash('Todos los documentos aprobados. El aprendiz pasa a estado PRE-MATRÍCULA', 'success')
        elif not todos_aprobados:
            docs_pendientes = [d for d in documentos_activos if d.validado != True]
            flash(f'Se aprobaron {documentos_aprobados} documentos. Aún faltan {len(docs_pendientes)} por aprobar.', 'info')
    else:
        flash('No hay documentos pendientes para aprobar', 'info')

    return redirect(url_for('docente.ver_matricula', matricula_id=matricula_id))

@docente_bp.route('/descargar-documentos-grupo/<int:grupo_id>')
@login_required
@docente_required
def descargar_documentos_grupo(grupo_id):
    """Descargar todos los documentos de un grupo en PDF unificado"""
    from app.services.formato_service import generar_pdf_unificado_grupo

    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()
    if not colegio:
        flash('No tiene un colegio asignado', 'danger')
        return redirect(url_for('docente.dashboard'))

    grupo = Grupo.query.get_or_404(grupo_id)

    # Obtener todos los aprendices del grupo que pertenecen al colegio del docente
    aprendices = Aprendiz.query.filter_by(grupo_id=grupo_id, colegio_id=colegio.id).all()

    if not aprendices:
        flash('No hay aprendices en este grupo', 'warning')
        return redirect(url_for('docente.matriculas'))

    try:
        # Generar PDF unificado
        pdf_path = generar_pdf_unificado_grupo(grupo, aprendices)

        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=os.path.basename(pdf_path)
        )

    except Exception as e:
        current_app.logger.error(f"Error al crear PDF unificado del grupo: {e}")
        flash(f'Error al crear PDF unificado: {str(e)}', 'danger')
        return redirect(url_for('docente.matriculas'))

@docente_bp.route('/descargar-documentos-aprendiz/<int:aprendiz_id>')
@login_required
@docente_required
def descargar_documentos_aprendiz(aprendiz_id):
    """Descargar todos los documentos de un aprendiz específico en PDF unificado"""
    from app.services.formato_service import generar_pdf_unificado_aprendiz

    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()
    aprendiz = Aprendiz.query.get_or_404(aprendiz_id)

    # Verificar que el aprendiz pertenece al colegio del docente
    if aprendiz.colegio_id != colegio.id:
        flash('No tiene permisos para acceder a este aprendiz', 'danger')
        return redirect(url_for('docente.dashboard'))

    matricula = Matricula.query.filter_by(aprendiz_id=aprendiz.id).first()
    if not matricula:
        flash('El aprendiz no tiene matrícula', 'warning')
        return redirect(url_for('docente.matriculas'))

    documentos = Documento.query.filter_by(matricula_id=matricula.id).filter(
        Documento.reemplazado_por == None
    ).all()

    if not documentos:
        flash('El aprendiz no tiene documentos para descargar', 'warning')
        return redirect(url_for('docente.ver_matricula', matricula_id=matricula.id))

    try:
        # Generar PDF unificado
        pdf_path = generar_pdf_unificado_aprendiz(aprendiz, documentos)

        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=os.path.basename(pdf_path)
        )

    except Exception as e:
        current_app.logger.error(f"Error al crear PDF unificado: {e}")
        flash(f'Error al crear PDF unificado: {str(e)}', 'danger')
        return redirect(url_for('docente.ver_matricula', matricula_id=matricula.id))


@docente_bp.route('/generar-sofia', methods=['GET', 'POST'])
@login_required
@docente_required
def generar_sofia():
    """Página para generar formato SOFIA Plus con filtros (limitado al colegio del docente)"""
    # Obtener colegio del docente
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No se encontró colegio asignado', 'danger')
        return redirect(url_for('docente.dashboard'))

    if request.method == 'GET':
        # Obtener opciones de filtro para el docente
        opciones = SofiaService.get_opciones_filtro_docente(colegio.id)
        return render_template('docente/generar_sofia.html', **opciones)

    # POST: Generar archivo
    filtro_tipo = request.form.get('filtro_tipo')
    filtro_id = request.form.get('filtro_id')

    if not filtro_tipo:
        flash('Debe seleccionar un tipo de filtro', 'warning')
        return redirect(url_for('docente.generar_sofia'))

    # Convertir filtro_id a entero si existe
    if filtro_id:
        try:
            filtro_id = int(filtro_id)
        except ValueError:
            flash('ID de filtro inválido', 'danger')
            return redirect(url_for('docente.generar_sofia'))

    # Generar formato (con restricción de colegio)
    success, message, file_path = SofiaService.generar_formato_sofia(
        filtro_tipo=filtro_tipo,
        filtro_id=filtro_id if filtro_tipo != 'colegio' else colegio.id,  # Forzar su colegio
        docente_colegio_id=colegio.id
    )

    if success:
        # Descargar archivo
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=os.path.basename(file_path)
        )
    else:
        flash(message, 'danger')
        return redirect(url_for('docente.generar_sofia'))


# ============================================
# GESTIÓN DE DOCUMENTOS SIMAT
# ============================================

@docente_bp.route('/simat')
@login_required
@docente_required
def simat():
    """Página de gestión de documentos SIMAT"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No tiene un colegio asignado', 'danger')
        return redirect(url_for('docente.dashboard'))

    # Obtener documentos SIMAT del colegio
    documentos = DocumentoSIMAT.query.filter_by(colegio_id=colegio.id).order_by(
        DocumentoSIMAT.created_at.desc()
    ).all()

    # Obtener grupos del colegio para el formulario
    grupos = Grupo.query.filter_by(colegio_id=colegio.id).order_by(Grupo.nombre).all()

    return render_template('docente/simat.html',
                         colegio=colegio,
                         documentos=documentos,
                         grupos=grupos)


@docente_bp.route('/simat/subir', methods=['POST'])
@login_required
@docente_required
def subir_simat():
    """Subir documento SIMAT"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No tiene un colegio asignado', 'danger')
        return redirect(url_for('docente.dashboard'))

    # Validar archivo
    if 'archivo' not in request.files:
        flash('No se seleccionó ningún archivo', 'warning')
        return redirect(url_for('docente.simat'))

    archivo = request.files['archivo']
    if archivo.filename == '':
        flash('No se seleccionó ningún archivo', 'warning')
        return redirect(url_for('docente.simat'))

    # Obtener tipo y grupo (si aplica)
    tipo = request.form.get('tipo', 'COLEGIO')
    grupo_id = request.form.get('grupo_id')

    # Validar que si es por grupo, se seleccionó un grupo
    if tipo == 'GRUPO' and not grupo_id:
        flash('Debe seleccionar un grupo', 'warning')
        return redirect(url_for('docente.simat'))

    # Validar extensión del archivo
    extensiones_permitidas = {'pdf', 'xlsx', 'xls', 'doc', 'docx'}
    if '.' not in archivo.filename or archivo.filename.rsplit('.', 1)[1].lower() not in extensiones_permitidas:
        flash('Tipo de archivo no permitido. Use PDF, Excel o Word', 'danger')
        return redirect(url_for('docente.simat'))

    try:
        # Generar nombre seguro para el archivo
        filename = secure_filename(archivo.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        extension = filename.rsplit('.', 1)[1].lower()
        nombre_archivo = f"SIMAT_{colegio.id}_{timestamp}.{extension}"

        # Crear directorio si no existe
        upload_dir = os.path.join(current_app.root_path, 'uploads', 'simat')
        os.makedirs(upload_dir, exist_ok=True)

        # Guardar archivo
        ruta_archivo = os.path.join(upload_dir, nombre_archivo)
        archivo.save(ruta_archivo)

        # Crear registro en base de datos
        documento_simat = DocumentoSIMAT(
            tipo=tipo,
            colegio_id=colegio.id,
            grupo_id=int(grupo_id) if grupo_id and tipo == 'GRUPO' else None,
            subido_por=current_user.id,
            ruta_archivo=ruta_archivo,
            nombre_archivo_original=archivo.filename,
            estado='PENDIENTE'
        )

        db.session.add(documento_simat)
        db.session.commit()

        flash('Documento SIMAT subido exitosamente. Pendiente de aprobación.', 'success')

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al subir documento SIMAT: {e}")
        flash(f'Error al subir documento: {str(e)}', 'danger')

    return redirect(url_for('docente.simat'))


@docente_bp.route('/simat/descargar/<int:simat_id>')
@login_required
@docente_required
def descargar_simat(simat_id):
    """Descargar documento SIMAT"""
    colegio = Colegio.query.filter_by(docente_enlace_id=current_user.id).first()

    if not colegio:
        flash('No tiene un colegio asignado', 'danger')
        return redirect(url_for('docente.dashboard'))

    # Verificar que el documento pertenece al colegio del docente
    documento = DocumentoSIMAT.query.filter_by(id=simat_id, colegio_id=colegio.id).first_or_404()

    if not os.path.exists(documento.ruta_archivo):
        flash('El archivo no existe', 'danger')
        return redirect(url_for('docente.simat'))

    return send_file(
        documento.ruta_archivo,
        as_attachment=True,
        download_name=documento.nombre_archivo_original
    )
